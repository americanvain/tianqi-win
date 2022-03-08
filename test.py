#-*- coding:utf-8 -*-

from openpyxl import load_workbook,Workbook
from datetime import datetime,timedelta

wb =load_workbook("weather_data.xlsx")
ws =wb.active
# print(ws.max_row)
# print(ws.cell(row=ws.max_row,column=1).value)
y=datetime.strptime(ws.cell(row=ws.max_row,column=1).value,'%Y-%m-%d %H:%M:%S')
print(y)
# # print(int(ws.cell(30,1).value))
date_str ='202203071900'
yy = datetime.strptime(date_str,'%Y%m%d%H%M')
print(yy.__gt__(y))
yyy=yy-y
# # if yyy.days<1:
# #     delt=yyy
# #     pass
print(yyy.days)
print(yyy.seconds/3600)
# timelist=[]
# print(y)
# timehaha=y-timedelta(hours=1)
# print(timehaha)
# timelist.append(timehaha)
# print(str(timelist[0]))
# colA = ws['A']
# for eachcell in colA:
#     print((eachcell.value))
# for index in range(2,-1,-1):
#     print(index)