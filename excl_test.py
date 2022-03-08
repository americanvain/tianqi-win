#-*- coding:utf-8 -*-

import os
from openpyxl import Workbook,load_workbook
from datetime import datetime,timedelta


def excel_creat():
    if not(os.path.exists("./original_data")):
        os.mkdir("original_data")
    if not(os.path.exists("./weather_data.xlsx")):
        wb = Workbook()
        ws =wb.active
        # cell = ws["A"]
        # cell[0].data_type='d'
        ws.append(["时间","温度","空气质量","相对湿度","风力"])
        wb.save("weather_data.xlsx")
        wb.close()
    
        
# def excel_process(shijian,tem,kongqi,shidu,fengli):
#     wb = load_workbook("weather_data.xlsx")
#     ws = wb.active
#     ws.append([shijian,tem,kongqi,shidu,fengli])
#     wb.save("weather_data.xlsx")
#     pass

# def excel_process(shijian=[],tem=[],kongqi=[],shidu=[],fengli=[]):
#     wb = load_workbook("weather_data.xlsx")
#     ws = wb.active
#     jutitime=datetime.strptime(shijian[-1],'%Y%m%d%H%M')
#     lasttime = datetime.strptime(ws.cell(row=ws.max_row,column=1).value,'%Y-%m-%d %H:%M:%S')
#     delta=lasttime-jutitime
#     cout
#     if delta.days<0:
#         cout = delta.seconds/3600
#     else:
#         cout=len(shijian)
#     for index in range(cout,0,-1):
#         ws.append([shijian[index],tem[index],kongqi[index],shidu[index],fengli[index]])
#     wb.save("weather_data.xlsx")
#     pass


def excel_process(jieguo_list,time_now):
    wb = load_workbook("weather_data.xlsx")
    ws = wb.active
    cout=len(jieguo_list)
    print(ws.max_row)
    time_now_get_last = time_now-timedelta(days=1)
    time_to_write=time_now_get_last
    if ws.max_row>1:
        lasttime = datetime.strptime(ws.cell(row=ws.max_row,column=1).value,'%Y-%m-%d %H:%M:%S')        
        if time_now_get_last.__ge__(lasttime):
            cout=len(jieguo_list)
            if time_now_get_last.__eq__(lasttime):
                ws.cell(row=ws.max_row,column=3,value=jieguo_list[25]['od27'])
                cout=cout-1
                pass
        else:
            delta=lasttime-time_now_get_last
            cout = len(jieguo_list)-delta.seconds/3600
    for index in range(cout,-1,-1):
        time_to_write=time_to_write+timedelta(hours=1)
        ws.append([str(time_to_write),jieguo_list[index]['od22'],jieguo_list[index]['od28'],jieguo_list[index]['od27'],jieguo_list[index]['od24']+jieguo_list[index]['od25']+'级'])

    wb.save("weather_data.xlsx")
    pass

# wb = Workbook()
# ws =wb.active
# cell = ws["A"]
# cell[0].data_type='d'
# print(cell[0].data_type)